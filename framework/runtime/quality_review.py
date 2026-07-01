#!/usr/bin/env python3
"""
quality_review.py — REVISAO HUMANA por capitulo, sem IA-julga-IA (piso de qualidade de verdade).

A back-translation (Opus julgando Sonnet/Haiku) custa e nao substitui um humano lendo o pt-BR. Aqui o
humano E o juiz: o `export` gera UM CSV com o CAPITULO INTEIRO (todas as linhas, p/ leitura integral),
mas cada linha vem MARCADA de forma 100% DETERMINISTICA (sem IA) com uma tag dizendo ONDE avaliar —
risco alto, amostra do tier barato, ou flags baratas (identico-a-fonte=provavel nao-traduzido, outlier
de tamanho, `largura`=segmento estoura o balao no jogo, marcador pt-PT). Linha sem tag = passa o olho;
tag preenchida = "avalie aqui".

O humano devolve o MESMO arquivo. A DECISAO e dele, numa coluna propria: escreve **CORRIGIR** (ou
corrigir) na coluna `marcar` da linha que quer mudar. O `apply` le SO as linhas marcadas CORRIGIR.
Na linha marcada, preenche UMA:
  - coluna `correcao` = o texto certo  -> aplico VERBATIM (zero IA: so gate de charset/round-trip + merge);
  - coluna `nota` (sem correcao) = instrucao (ex.: "encurtar", "tom formal") -> IA re-traduz SO aquela
    linha seguindo a nota (cirurgico, nunca a cena inteira).
Linha SEM CORRIGIR = aprovada, nao toco. A coluna `revisar` e so DICA de onde olhar (deterministica +
o micro-QA da IA JA pago — verdict 'revise' da back-translation); NAO decide nada, $0, sem nova API.

O `apply` processa EXATAMENTE o que foi marcado e re-verifica round-trip dos capitulos tocados. Governanca:
HUMANO propoe -> gate (charset/paridade/round-trip) aprova -> script aplica. Sem work-text no .py.

Uso:
  python quality_review.py export <projeto> [<cap>] [--csv]   # XLSX amigavel (default); omita cap = JOGO TODO
  python quality_review.py apply  <projeto> <arquivo-devolvido>   # le XLSX ou CSV
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))
import artifact_io  # noqa: E402  (leitura compartilhada: scenes/translations_map)
import context_pack  # noqa: E402
import model  # noqa: E402
import paths  # noqa: E402

# 'marcar' = a COLUNA DE DECISÃO do humano: ele escreve CORRIGIR (ou corrigir) na linha que quer mudar.
# O apply lê SÓ as linhas marcadas. 'revisar' é só DICA ($0, micro-QA da IA) de onde olhar — não decide nada.
COLS = ["scene", "offset", "speaker", "risk", "revisar", "source_en", "target_pt",
        "caixa", "marcar", "correcao", "nota", "repete"]

# Marcadores pt-PT de ALTA precisao (raros no pt-BR falado) — heuristica, por isso a tag leva '?'.
_PTPT = re.compile(r"\b(tens|estás|fazes|podes|queres|deves|vês|hás)\b|\btem de\b|\bhás de\b", re.IGNORECASE)

# Expressoes culturais/idiomaticas EN (americana/britanica) na FONTE — sinal p/ o revisor checar se a
# localizacao pt-BR ficou equivalente (nao so traduzida literalmente). Deterministico, $0, varre o source.
# Alta precisao: preferir expressoes multi-palavra inequivocas; evitar palavras comuns isoladas.
_CULTURAL_EN = re.compile(
    r"\b("
    # promessas / juramentos
    r"cross my heart( and hope to die)?"
    r"|pinky (promise|swear)|scout'?s honor"
    # vida / morte / sorte
    r"|bite the bullet|kick the bucket|bite the dust"
    r"|on (your|my|his|her|their|one'?s) last legs"
    r"|nine lives|cat'?s got (your|my|his|her|their) tongue"
    r"|curiosity killed the cat"
    r"|every dog has (its|his) day"
    r"|when pigs fly|raining cats and dogs"
    r"|once in a blue moon|a blessing in disguise"
    # sucesso / fracasso
    r"|break a leg|hit the jackpot|paint the town red"
    r"|hit (the ground running|rock bottom|the nail on the head)"
    r"|by the skin of (my|your|his|her|their|one'?s) teeth"
    r"|pass with flying colors"
    r"|miss the boat|drop the ball|blow (it|your|my|his|her|their) chance"
    r"|throw in the towel|bite off more than (you|he|she|they|one) can chew"
    # facilidade / dificuldade
    r"|piece of cake|not rocket science|it'?s? no (picnic|walk in the park)"
    r"|easier said than done|uphill battle"
    r"|a tough nut to crack|no (pain[,]? no gain|strings attached)"
    # honestidade / segredo
    r"|let the cat out of the bag|spill the beans"
    r"|straight from the horse'?s mouth|read between the lines"
    r"|cards on the table|skeleton in the (closet|cupboard)"
    r"|elephant in the room|tip of the iceberg"
    # traição / lealdade
    r"|throw (someone )?under the bus|stab (someone )?in the back"
    r"|bite the hand that feeds (you|him|her|them|one)"
    r"|wolf in sheep'?s clothing|turn a blind eye"
    r"|bury the hatchet|burn (your|my|our|their|one'?s|the) bridges?"
    r"|leave (someone )?in the lurch|give (someone )?the cold shoulder"
    # dinheiro / custo
    r"|cost(s)? an arm and a leg|cost a pretty penny"
    r"|bring home the bacon|make ends meet"
    r"|worth (its|his|her|their|your|my|one'?s) weight in gold"
    r"|dime a dozen|a penny for your thoughts"
    # tempo / urgência
    r"|burn the midnight oil|against the clock"
    r"|at the drop of a hat|jump the gun"
    r"|better late than never|time flies"
    r"|call it a day|back to (the drawing board|square one)"
    # confusão / erro
    r"|barking up the wrong tree|put (your|my|his|her|their|one'?s) foot in (your|my|his|her|their|one'?s) mouth"
    r"|make a mountain out of a molehill|a storm in a (teacup|teapot)"
    r"|lose (your|my|his|her|their|one'?s) marbles|go off the rails"
    r"|beat around the bush|miss the point"
    r"|get out of hand|in over (your|my|his|her|their|one'?s) head"
    r"|don'?t count your chickens( before they hatch)?"
    r"|don'?t cry over spilled milk|out of the frying pan (and )?into the fire"
    # aparência / julgamento
    r"|don'?t judge a book by its cover|all bark (and )?no bite"
    r"|actions speak louder than words|the pot calling the kettle black"
    r"|two peas in a pod|blood is thicker than water"
    # pressão / persistência
    r"|bend over backwards|go the extra mile|go above and beyond"
    r"|hold your horses|keep your nose to the grindstone"
    r"|hang in there|keep your chin up|keep (your|my|his|her|their|one'?s) head above water"
    r"|add fuel to (the )?fire|add insult to injury"
    r"|twist (someone'?s|your|my|his|her|their|one'?s) arm"
    # posição / indecisão
    r"|on the fence|sit on the fence|up in the air"
    r"|on thin ice|in hot water|in (deep|hot) water"
    r"|the ball is in your court|on cloud nine"
    # relacionamento / social
    r"|break the ice|see eye to eye|head over heels"
    r"|pull (someone'?s|your|my|his|her|their) leg"
    r"|steal (someone'?s|your|my|his|her|their) thunder"
    r"|rule of thumb|the best of both worlds"
    r"|cut corners|kill two birds with one stone"
    r"|(touch|knock on) wood|come in handy"
    r"|devil'?s advocate|double.edged sword"
    r"|think outside (the|of the) box|cut to the chase"
    r"|let sleeping dogs lie|don'?t push (your|my|his|her|their|one'?s) luck"
    r"|look on the bright side|every cloud has a silver lining"
    r"|the last straw|the tip of the iceberg"
    r"|jump on the bandwagon|ring a bell"
    r"|get cold feet|in the same boat"
    r"|pull (yourself|oneself) up by (your|the) bootstraps"
    r"|the plot thickens|a red herring"
    r"|cast (the )?pearls? before swine|a wolf at the door"
    r"|once bitten[,]? twice shy|burn your (fingers|hands)"
    r"|get a taste of (your|my|his|her|their|one'?s) own medicine"
    r"|it takes two to tango|you can'?t have your cake and eat it( too)?"
    r"|actions speak louder than words|get the ball rolling"
    r"|face the music|out of the blue"
    # girias americanas
    r"|holy (cow|moly|smokes|guacamole|mackerel)|what the heck|what in tarnation"
    r"|y'all|ya'?ll|you bet|good grief|for crying out loud"
    r"|cut me some slack|no sweat|piece of my mind"
    r"|give (me|him|her|them|you|us|someone) a (break|hand|piece of)"
    # marcadores britanicos
    r"|blimey|crikey|bloody hell|cheerio|innit"
    r"|gobsmacked|knackered|taking the mickey|bob'?s your uncle"
    r"|cor blimey|gutted\b|chuffed|(bang|spot) on|chin up"
    r")\b",
    re.IGNORECASE
)

# Largura do BALAO: o byte_budget garante que cabe no ARQUIVO (reinsercao), NAO que cabe na largura
# VISUAL do balao. Cada segmento entre tokens de quebra (`\n`) e UMA linha exibida.
# TESTE SEM IN-GAME (deterministico): a RE da fonte do jogo (Font.fnt) mostrou um ATLAS DE GRADE
# UNIFORME (celula fixa por glifo, zero largura proporcional) -> a fonte e MONOESPACADA, logo
# nº de caracteres VISIVEIS == largura real em pixels. Calibrado pelo corpus EN (que ja roda no jogo):
# o ENVELOPE = o maior segmento de DIALOGO que o EN ja renderizou (provado em tela). A massa fica <=54,
# mas ha linhas reais ate ~64 que o jogo embarcou -> usar o MAXIMO provado (62, abaixo do maior real),
# nao o 99-pct (54 super-marcaria linhas que de fato cabem). Segmento pt-BR > 62 = mais largo que
# qualquer EN que o jogo mostrou -> risco de sair do balao. (Narracao/credito em caixa larga pode passar
# disso no EN; sao poucos e fora do balao de dialogo.)
# IMPORTANTE: medir SO os glifos visiveis — tokens `{c5}/{c-1}/{W12}/...` NAO renderizam largura
# (ex.: "H{W10}A{W10}V{W10}E..." sao poucos chars na tela). Ver _VISIBLE_RX.
WIDTH_MAX = 62
_VISIBLE_RX = re.compile(r"\{c-?\d*\}|\{W\d+\}|\{COLOR\}|\{END\}")   # tokens que NAO ocupam largura visual


# BoF4: detecta codigos de controle do DAT no formato [XX] (hex 1-2 chars)
_BOF4_HEX_RX = re.compile(r"\[[0-9A-Fa-f]{1,2}\]")
# par: codigo de controle + byte de face (@=0x40, A=0x41, B=0x42 ... G=0x47)
_BOF4_FACE_PAIR_RX = re.compile(r"\[[0-9A-Fa-f]{1,2}\][@A-G]")


def _display_text(s: str) -> str:
    """Limpa codigos de controle BoF4 para exibicao no Excel (nao altera dados armazenados).
    Detectado automaticamente pela presenca de [XX] — nao precisa de config por projeto.
    [01] -> newline, [02] -> ' // ' (page break), demais codigos e bytes-de-face removidos."""
    if not s or not _BOF4_HEX_RX.search(s):
        return s
    s = s.replace("[01]", "\n").replace("[02]", " // ")
    s = _BOF4_FACE_PAIR_RX.sub("", s)           # remove [XX] + byte de face juntos
    s = _BOF4_HEX_RX.sub("", s)                 # remove codigos remanescentes
    s = re.sub(r"[ \t]+", " ", s)               # normaliza espacos (sem destruir \n)
    s = re.sub(r"(^\s*(//\s*)+|(//\s*)*\s*$)", "", s)  # strip page-break markers nas bordas
    return s.strip()


def _norm_cmp(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").replace("\\n", " ")).strip().lower()


def _flags(source: str, target: str, risk: str, sampled: bool, bt_revise: bool = False) -> str:
    """Tags de 'onde avaliar' — DICA p/ o humano, NUNCA decisao (a decisao e a coluna CORRIGIR).
    Tudo $0: deterministico OU reusa o micro-QA da IA JA pago (verdict 'revise' da back-translation).
    '' = linha sem sinal (passa o olho)."""
    fl = []
    if bt_revise:
        fl.append("micro-qa:revise")                      # a back-translation (IA, ja paga) achou divergencia
    if risk in ("high", "critical"):
        fl.append(f"risco:{risk}")
    if sampled:
        fl.append("amostra")
    if target and _norm_cmp(source) == _norm_cmp(target):
        fl.append("identico-fonte")                       # provavel nao-traduzido
    slen, tlen = model._translit_len(source), model._translit_len(target)
    if slen and (tlen > slen * 3 or (slen > 8 and tlen < slen * 0.4)):
        fl.append("tamanho")                              # outlier de comprimento
    if any(model._translit_len(_VISIBLE_RX.sub("", seg)) > WIDTH_MAX     # so glifos VISIVEIS (sem tokens)
           for seg in re.split(re.escape(context_pack.TOKEN) + r"|\\n|\n| // ", target or "")):
        fl.append("largura")                              # segmento estoura a largura do balao (in-game)
    if _PTPT.search(target or ""):
        fl.append("pt-PT?")
    if _CULTURAL_EN.search(source or ""):
        fl.append("expr-cultural")   # idioma EN na fonte — checar se localizacao PT-BR ficou equivalente
    return ";".join(fl)


def _envelope(s):
    """(maior largura de segmento VISIVEL, nº de segmentos) de uma string. Quebra por TOKEN (\\n) OU
    '\\n' literal; desconta tokens de formatacao (nao renderizam). Monospace -> chars = pixels."""
    parts = re.split(re.escape(context_pack.TOKEN) + r"|\\n|\n| // ", s or "")
    widths = [model._translit_len(_VISIBLE_RX.sub("", p)) for p in parts]
    nseg = sum(1 for p in parts if p.strip())
    return (max(widths) if widths else 0), max(nseg, 1)


def _box_verdict(source: str, target: str) -> str:
    """Veredito DETERMINISTICO de caixa, por linha: compara o pt-BR com a SUA EN (mesmo offset = MESMA
    caixa). Como char=px (fonte monospace) e a EN JA rodou no jogo:
      - pt-BR <= EN (largura E nº linhas) -> '' (CABE, provado; a caixa ja exibiu esse envelope).
      - pt-BR > EN mas <= maior dialogo EN do jogo -> 'rever +Nc/+NL' (pode usar a folga da caixa).
      - pt-BR > maior dialogo EN do jogo -> 'ESTOUROU +Nc' (mais largo que tudo que o jogo mostrou)."""
    ew, el = _envelope(source)
    pw, pl = _envelope(target)
    if pw <= ew and pl <= el:
        return ""
    delta = []
    if pw > ew:
        delta.append(f"+{pw - ew}c")
    if pl > el:
        delta.append(f"+{pl - el}L")
    head = "ESTOUROU" if pw > WIDTH_MAX else "rever"
    return f"{head} {' '.join(delta)}".strip()


def export(root, chapter) -> list[dict]:
    """CSV-rows do capitulo inteiro, cada linha marcada. Determinista (sem rede)."""
    root = Path(root)
    rows = []
    for scene in artifact_io.scenes(root, chapter):
        plan_lines = model._plan_lines(root, scene)
        if not plan_lines:
            continue
        tmap = artifact_io.translations_map(root, scene)
        sampled = {x["offset"] for x in model.sample_low_risk_lines(root, scene)}
        revise = _bt_revise_offsets(root, scene)          # micro-QA da IA JA pago ($0 ler)
        for ln in plan_lines:
            off = ln.get("offset", "")
            src = _display_text(ln.get("text_source", ""))
            tgt = _display_text(
                (tmap.get(off) or {}).get("t", "") if isinstance(tmap.get(off), dict)
                else ln.get("base_translation", "")
            )
            risk = ln.get("risk_level", "")
            rows.append({"scene": scene, "offset": off, "speaker": ln.get("speaker", ""),
                         "risk": risk,
                         "revisar": _flags(src, tgt, risk, off in sampled, off in revise),
                         "source_en": src, "target_pt": tgt,
                         "caixa": _box_verdict(src, tgt),
                         "marcar": "", "correcao": "", "nota": "", "repete": 1})
    pair_cnt = Counter((r["source_en"], r["target_pt"]) for r in rows)
    for r in rows:
        r["repete"] = pair_cnt[(r["source_en"], r["target_pt"])]
    return rows


def _bt_revise_offsets(root, scene) -> set:
    """Offsets que a back-translation (micro-QA da IA, JA paga) marcou 'revise'. So leitura ($0)."""
    sid = context_pack.scene_id_of(scene)
    btf = paths.back_translation(root, scene, sid)
    if not btf.is_file():
        return set()
    try:
        data = json.loads(btf.read_text(encoding="utf-8"))
    except Exception:
        return set()
    # stale=True: linha foi corrigida verbatim DEPOIS do bt -> bt antigo nao vale mais (nao emitir micro-qa)
    return {e.get("offset") for e in data.get("entries", [])
            if e.get("verdict") == "revise" and not e.get("stale")}


def width_violations(root, chapter=None) -> list:
    """GATE de risco ALTO: linhas 'ESTOUROU' — pt-BR mais largo que QUALQUER dialogo EN que o jogo ja
    mostrou (veredito determinístico da coluna 'caixa'; char=px monospace). Quebra quase certa -> loop ate
    zerar. ('rever' = cresceu vs a propria EN mas dentro do envelope = zona cinza, NAO gateia; '' = cabe,
    provado.) Ver _box_verdict."""
    return [r for r in export(root, chapter) if (r.get("caixa") or "").startswith("ESTOUROU")]


def write_csv(rows, out_path):
    with Path(out_path).open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in COLS})


# rotulos amigaveis (PT) p/ o XLSX, na MESMA ordem de COLS (a leitura mapeia por posicao)
_XLSX_HEAD = ["Cena", "Offset", "Falante", "Risco", "Revisar (onde olhar)", "Ingles (fonte)",
              "Portugues (atual)", "Caixa (cresceu vs EN?)", "Corrigir? (escreva CORRIGIR)",
              "Correcao (texto certo)", "Nota (instrucao p/ IA)", "Repete (N vezes no jogo)"]
# severidade -> cor da linha (a 1a tag presente vence; ordem = mais grave primeiro)
_XLSX_SEV = [("micro-qa", "F4CCCC"), ("critical", "FFC7CE"), ("high", "FFE2C7"),
             ("expr-cultural", "D5F5E3"),
             ("largura", "CFE2FF"), ("identico-fonte", "E8E8E8"),
             ("tamanho", "FFF0C7"), ("pt-PT", "EAD9F2")]
_XLSX_INPUT = "FFF7CC"   # amarelo claro nas colunas de input do humano (Corrigir?/Correcao/Nota)


def write_xlsx(rows, out_path):
    """Relatorio AMIGAVEL p/ o revisor humano (Excel/LibreOffice): aba 'Leia-me' (instrucoes+legenda+
    contagem) + aba 'Revisao' com cabecalho congelado, autofiltro, cor por tipo de erro, colunas de
    input em amarelo e EN/PT com quebra de linha. O `apply` le este xlsx de volta (mapeado por posicao)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("o relatorio XLSX amigavel requer 'openpyxl' (pip install openpyxl). "
                           "Ou use --csv p/ o CSV cru.") from e

    cnt = Counter(t for r in rows if r.get("revisar") for t in r["revisar"].split(";"))
    marked = sum(1 for r in rows if r.get("revisar"))
    wb = Workbook()

    # ── aba Leia-me ──────────────────────────────────────────────────────────────
    intro = wb.active
    intro.title = "Leia-me"
    intro.column_dimensions["A"].width = 24
    intro.column_dimensions["B"].width = 68
    intro.sheet_view.showGridLines = False

    def _c(row, col, value="", bold=False, size=10, color="1A1A1A",
           bg=None, wrap=False, halign="left", valign="center"):
        cell = intro.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        return cell

    def _sec(row, label, bg="1F3864"):
        _c(row, 1, "  " + label, bold=True, size=11, color="FFFFFF", bg=bg, valign="center")
        _c(row, 2, bg=bg)
        intro.row_dimensions[row].height = 24

    def _step(row, num, text):
        _c(row, 1, str(num) + ".", bold=True, size=11, color="1F3864",
           bg="D6E4F7", halign="center", valign="top")
        _c(row, 2, text, size=10, wrap=True, valign="top")
        intro.row_dimensions[row].height = max(18, min(60, 15 + len(text) // 5))

    def _leg(row, bg_hex, label, desc):
        _c(row, 1, label, bold=True, size=9, color="333333",
           bg=bg_hex, halign="center", valign="center")
        _c(row, 2, desc, size=9, wrap=True, valign="top")
        intro.row_dimensions[row].height = 24

    def _kv(row, key, val, key_bold=True):
        _c(row, 1, key, bold=key_bold, size=10, color="333333")
        _c(row, 2, val, size=10, color="333333")

    r = 1
    # Titulo
    _c(r, 1, "GUIA DO REVISOR", bold=True, size=16, color="FFFFFF",
       bg="1F3864", halign="center", valign="center")
    _c(r, 2, bg="1F3864")
    intro.row_dimensions[r].height = 36
    r += 1; intro.row_dimensions[r].height = 6; r += 1

    # Secao: como usar
    _sec(r, "COMO USAR"); r += 1
    _step(r, 1, "Va para a aba 'Revisao'. Voce vera todas as linhas do jogo, uma por linha."); r += 1
    _step(r, 2, "A coluna 'Revisar (onde olhar)' e so DICA — indica onde o sistema ou o micro-QA "
                "da IA apontou algo. Filtre por nao-vazias para priorizar. NAO decide nada."); r += 1
    _step(r, 3, "VOCE decide: na coluna 'Corrigir? (escreva CORRIGIR)' escreva CORRIGIR na linha "
                "que quer mudar. So as linhas marcadas serao processadas."); r += 1
    _step(r, 4, "Na MESMA linha marcada, preencha UMA das duas colunas:\n"
                "  'Correcao' = o texto EXATO que deve ir para o jogo (verbatim, $0, sem IA);\n"
                "  'Nota' = instrucao para a IA reescrever so aquela linha "
                "(ex.: 'encurtar', 'mais formal', 'tom brincalhao')."); r += 1
    _step(r, 5, "Linha boa = deixe em branco (nao escreva CORRIGIR). "
                "Salve e devolva o arquivo na pasta devolvido/."); r += 1
    r += 1; intro.row_dimensions[r].height = 6; r += 1

    # Secao: legenda
    _sec(r, "LEGENDA DAS CORES  (coluna Revisar)"); r += 1
    _leg(r, "F4CCCC", "micro-qa:revise",
         "A back-translation da IA achou divergencia de sentido — releia com atencao"); r += 1
    _leg(r, "FFC7CE", "critical",
         "Linha de risco CRITICO (voz, sentido central, spoiler) — leitura obrigatoria"); r += 1
    _leg(r, "FFE2C7", "high",
         "Linha de risco ALTO — leia com cuidado"); r += 1
    _leg(r, "D5F5E3", "expr-cultural",
         "Expressao idiomatica EN detectada na fonte (ex.: 'cross my heart', 'break a leg', 'spill the beans') — "
         "verifique se a localizacao PT-BR ficou equivalente e natural, nao so traduzida ao pe da letra"); r += 1
    _leg(r, "CFE2FF", "largura",
         "Texto pode SAIR do balao no jogo — encurte se necessario"); r += 1
    _leg(r, "E8E8E8", "identico-fonte",
         "Igual ao ingles — provavel nao-traduzido (SFX/rotulo pode ficar igual; confirme)"); r += 1
    _leg(r, "FFF0C7", "tamanho",
         "Traducao muito mais longa ou curta que o original"); r += 1
    _leg(r, "EAD9F2", "pt-PT?",
         "Marcador de portugues de Portugal detectado (tens/estás/podes...) — adaptar para pt-BR"); r += 1
    r += 1; intro.row_dimensions[r].height = 6; r += 1

    # Secao: legenda coluna CAIXA
    _sec(r, "LEGENDA DA COLUNA 'CAIXA (cresceu vs EN?)'"); r += 1
    _leg(r, "FFFFFF", "(vazio)",
         "CABE: o envelope pt-BR e <= ingles — a mesma caixa ja exibiu esse texto em jogo"); r += 1
    _leg(r, "FFE2C7", "rever +Nc",
         "Traducao N caracteres MAIS LARGA que o ingles nessa linha (mas ainda dentro do maior "
         "dialogo EN do jogo) — PODE caber usando a folga da caixa; verifique em jogo antes de encurtar"); r += 1
    _leg(r, "FFE2C7", "rever +NL",
         "Traducao tem N linhas a mais que o ingles — mesma logica; verifique em jogo"); r += 1
    _leg(r, "FFC7CE", "ESTOUROU +Nc",
         "Traducao N caracteres mais larga que o MAIOR dialogo ingles ja mostrado no jogo (limite: "
         f"{WIDTH_MAX} chars) — provavelmente sai da caixa; encurte e re-exporte para confirmar"); r += 1
    r += 1; intro.row_dimensions[r].height = 6; r += 1

    # Secao: legenda coluna REPETE
    _sec(r, "LEGENDA DA COLUNA 'REPETE (N vezes no jogo)'"); r += 1
    _leg(r, "FFFFFF", "1",
         "Esta linha e unica — revise normalmente"); r += 1
    _leg(r, "E8D5F5", "> 1",
         "Este par (ingles + portugues) aparece N vezes no jogo (NPCs com mesmo dialogo, menus "
         "repetidos, etc.). Basta revisar UMA ocorrencia — se precisar corrigir, marque CORRIGIR "
         "em TODAS as linhas com o mesmo texto (cada offset e um ponto de reinsercao independente)."); r += 1
    r += 1; intro.row_dimensions[r].height = 6; r += 1

    # Secao: resumo
    _sec(r, "RESUMO DO ARQUIVO"); r += 1
    _kv(r, "Total de linhas", len(rows)); r += 1
    _kv(r, "Com dica para avaliar", f"{marked}  ({round(100*marked/len(rows)) if rows else 0}%)"); r += 1
    if cnt:
        r += 1
        _c(r, 1, "Distribuicao das dicas:", bold=True, size=9, color="555555"); r += 1
        for tag, count in cnt.most_common():
            _c(r, 1, f"    {tag}", size=9, color="555555")
            _c(r, 2, count, size=9, color="555555"); r += 1

    ws = wb.create_sheet("Revisao")
    ws.append(_XLSX_HEAD)
    hfill = PatternFill("solid", fgColor="2F5496")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")
    inputfill = PatternFill("solid", fgColor=_XLSX_INPUT)
    cx = COLS.index("caixa") + 1                           # coluna 'caixa' (1-indexed)
    rx = COLS.index("repete") + 1                          # coluna 'repete' (1-indexed)
    box_red = PatternFill("solid", fgColor="FFC7CE")       # ESTOUROU (passou do maior EN)
    box_org = PatternFill("solid", fgColor="FFE2C7")       # rever (cresceu vs EN)
    rep_fill = PatternFill("solid", fgColor="E8D5F5")      # repete > 1
    for r in rows:
        ws.append([r.get(c, "") for c in COLS])
        i = ws.max_row
        rev = r.get("revisar", "")
        fill = next((PatternFill("solid", fgColor=clr) for tag, clr in _XLSX_SEV if tag in rev), None)
        for col in range(1, len(COLS) + 1):
            cell = ws.cell(row=i, column=col)
            # openpyxl promove strings com '=' para data_type='f' (formula); quotePrefix=True
            # instrui o Excel a tratar o conteudo como texto simples (invisivel ao usuario).
            if isinstance(cell.value, str) and cell.value and cell.value[0] in "=+-@":
                cell.quotePrefix = True
            cell.font = Font(name="Arial", size=10)
            cell.alignment = wrap if col in (6, 7, 10, 11) else top
            if col in (9, 10, 11):                          # Corrigir?/Correcao/Nota = input do humano (amarelo)
                cell.fill = inputfill
            elif col == cx:                                 # 'caixa': cor por veredito (det.)
                cv = r.get("caixa", "")
                if cv.startswith("ESTOUROU"):
                    cell.fill = box_red
                elif cv:
                    cell.fill = box_org
            elif col == rx:                                 # 'repete': destaque se > 1
                if (r.get("repete") or 1) > 1:
                    cell.fill = rep_fill
            elif fill is not None:
                cell.fill = fill
    widths = {1: 10, 2: 11, 3: 14, 4: 9, 5: 22, 6: 50, 7: 50, 8: 18, 9: 16, 10: 42, 11: 28, 12: 12}
    for col, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = "A2"                                 # cabecalho fixo ao rolar
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLS)).column_letter}{ws.max_row}"
    wb.save(out_path)


_MAX_XLSX_MB = 20


def _read_xlsx_rows(path):
    """Le a aba 'Revisao' do xlsx devolvido -> lista de dicts {COLS: valor} (mapeado por POSICAO)."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("ler XLSX devolvido requer 'openpyxl' (pip install openpyxl).") from e
    size_mb = Path(path).stat().st_size / (1024 * 1024)
    if size_mb > _MAX_XLSX_MB:
        raise ValueError(f"XLSX muito grande ({size_mb:.1f} MB > {_MAX_XLSX_MB} MB); "
                         f"verifique se o arquivo e valido.")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Revisao"] if "Revisao" in wb.sheetnames else wb[wb.sheetnames[-1]]
    out, first = [], True
    for row in ws.iter_rows(values_only=True):
        if first:                                          # pula cabecalho
            first = False
            continue
        out.append({COLS[i]: ("" if i >= len(row) or row[i] is None else str(row[i]))
                    for i in range(len(COLS))})
    return out


def read_returned(path) -> dict:
    """Le o CSV ou XLSX devolvido -> {scene: {'verbatim': [(offset, texto)], 'nota': [(offset, instrucao)]}}.
    So entram as linhas que o HUMANO marcou explicitamente com CORRIGIR (coluna 'marcar', case-insensitive)
    E que tenham correcao OU nota. Linha sem CORRIGIR = aprovada, ignorada (le-se so o que o humano marcou).
    O campo '_total_marked' (chave interna) conta linhas marcadas CORRIGIR antes de qualquer guard."""
    p = Path(path)
    if p.suffix.lower() == ".xlsx":
        records = _read_xlsx_rows(p)
    else:
        with p.open(encoding="utf-8-sig", newline="") as fh:
            records = list(csv.DictReader(fh))
    by_scene: dict = {}
    total_marked = 0
    for r in records:
        scene, off = (r.get("scene") or "").strip(), (r.get("offset") or "").strip()
        marcar = (r.get("marcar") or "").strip().lower()
        cor, nota = (r.get("correcao") or "").strip(), (r.get("nota") or "").strip()
        if not scene or not off or marcar != "corrigir":
            continue                                       # so processa o que foi MARCADO 'corrigir'
        total_marked += 1
        # Guard path traversal: scene/off vêm de XLSX externo; sem separadores ou '..'
        if any(c in scene for c in ("/", "\\", "..")) or any(c in off for c in ("/", "\\")):
            continue
        if not cor and not nota:
            continue                                       # marcada mas sem texto/instrucao -> nada a aplicar
        slot = by_scene.setdefault(scene, {"verbatim": [], "nota": []})
        if cor:
            slot["verbatim"].append((off, cor))           # correcao verbatim vence a nota
        else:
            slot["nota"].append((off, nota))
    by_scene["_total_marked"] = total_marked              # metadado interno, removido pelo apply antes de iterar
    return by_scene


def _apply_verbatim(root, scene, pairs) -> int:
    """Grava o texto do humano em translations + plan (parity-fit; SEM IA). Retorna nº de linhas."""
    import json
    sid = context_pack.scene_id_of(scene)
    tf, pf = paths.translations(root, scene, sid), paths.translation_plan(root, scene, sid)
    if not tf.is_file() or not pf.is_file():
        return 0
    tdata = json.loads(tf.read_text(encoding="utf-8"))
    pdata = json.loads(pf.read_text(encoding="utf-8"))
    srcmap = {ln.get("offset", ""): ln.get("text_source", "") for ln in pdata.get("lines", [])}
    n = 0
    for off, txt in pairs:
        fitted = model._parity_fit(srcmap.get(off, ""), model._norm_t(txt))
        tdata.setdefault("lines", {}).setdefault(off, {})["t"] = fitted
        for ln in pdata.get("lines", []):
            if ln.get("offset") == off:
                ln["base_translation"] = fitted
        n += 1
    tf.write_text(json.dumps(tdata, ensure_ascii=False, indent=2), encoding="utf-8")
    pf.write_text(json.dumps(pdata, ensure_ascii=False, indent=2), encoding="utf-8")
    model.invalidate_back_translation(root, scene, [o for o, _ in pairs])  # crivo antigo nao vale mais
    return n


def returned_files(root, arg) -> list[Path]:
    """Resolve de ONDE ler a revisao devolvida: arquivo explicito, ou uma PASTA (todos os .xlsx/.csv
    dentro), ou — se nada for passado — o INBOX padrao (artifacts/qa_revisao/devolvido/). Ignora
    arquivos temporarios do Excel (~$...). Devolve [] se nao houver nada (apply vira no-op seguro)."""
    if arg:
        p = Path(arg)
        if p.is_file():
            return [p]
        if p.is_dir():
            base = p
        else:
            return []                                       # caminho dado mas inexistente
    else:
        base = paths.qa_inbox(root)
    if not base.is_dir():
        return []
    return sorted(f for f in base.iterdir()
                  if f.suffix.lower() in (".xlsx", ".csv") and not f.name.startswith("~$"))


# ----------------------------- TESTER in-game (print -> texto -> linha) -----------------------------
# O tester NAO le offset (in-game so ve TEXTO). Ele digita um trecho do pt-BR que apareceu na tela +
# larga o print (prova). Este localizador casa o trecho contra os approved_*.csv e devolve a linha.
# DETERMINISTICO ($0, sem IA, sem OCR): "transformar print em texto" = o olho do tester; o print e prova.
# O texto in-game e TRANSLITERADO (sem acento) -> dobramos acento dos dois lados pra casar (NFD).

def _fold(s: str) -> str:
    """Normaliza p/ casar com o que aparece NA TELA: tira tokens, dobra acento (NFD), minuscula, colapsa
    espacos. 'está tão' -> 'esta tao' (= o transliterado in-game)."""
    s = _VISIBLE_RX.sub("", (s or "").replace("\\n", " ").replace(context_pack.TOKEN, " "))
    s = "".join(c for c in unicodedata.normalize("NFD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _approved_index(root):
    """[(scene, offset, target_pt)] de TODAS as cenas (le approved_*.csv via artifact_io)."""
    out = []
    for scene in artifact_io.scenes(root, None):
        tmap = artifact_io.translations_map(root, scene)
        for off, v in (tmap or {}).items():
            t = v.get("t", "") if isinstance(v, dict) else ""
            if t:
                out.append((scene, off, t))
    return out


def locate(index, snippet):
    """Linhas cujo pt-BR (dobrado) CONTEM o trecho (dobrado). Trecho < 3 chars uteis -> [] (ambiguo demais)."""
    q = _fold(snippet)
    if len(q) < 3:
        return []
    return [(sc, off, t) for sc, off, t in index if q in _fold(t)]


def tester_to_review(root, relato_path):
    """Le relato_tester.csv (print, texto_visto, problema, sugestao) -> resolve cada relato p/ cena+offset
    e monta linhas CORRIGIR (mesmo formato do devolvido). Retorna (rows, ambiguos, nao_achados)."""
    root = Path(root)
    with Path(relato_path).open(encoding="utf-8-sig", newline="") as fh:
        relatos = list(csv.DictReader(fh))
    index = _approved_index(root)
    rows, ambiguous, missing = [], [], []
    for r in relatos:
        snip = (r.get("texto_visto") or "").strip()
        if not snip:
            continue
        prob = (r.get("problema") or "").strip()
        sug = (r.get("sugestao") or "").strip()
        prnt = (r.get("print") or "").strip()
        hits = locate(index, snip)
        if len(hits) == 1:
            sc, off, t = hits[0]
            rows.append({"scene": sc, "offset": off, "source_en": "", "target_pt": t,
                         "marcar": "CORRIGIR",
                         "correcao": sug,                                   # sugestao = verbatim ($0)
                         "nota": "" if sug else (f"{prob} (ver print {prnt})".strip())})  # senao = IA
        elif not hits:
            missing.append({"print": prnt, "texto_visto": snip, "problema": prob})
        else:
            ambiguous.append({"print": prnt, "texto_visto": snip, "n": len(hits),
                              "candidatos": [f"{sc}:{off}" for sc, off, _ in hits[:8]]})
    return rows, ambiguous, missing


def apply(root, csv_path, *, model_name=None, max_usd=None) -> dict:
    """Processa EXATAMENTE o devolvido: verbatim (0 IA) + nota (IA cirurgica por linha). `max_usd` so
    limita o caminho de IA (verbatim e sempre $0). Retorna {verbatim, ai, scenes, cost_usd,
    scenes_touched[], stopped_budget, total_marked, effectiveness_rate}.
    Persiste um registro em artifacts/qa_effectiveness.jsonl para rastrear o ciclo ao longo do tempo."""
    import time as _time
    root = Path(root)
    returned = read_returned(csv_path)
    total_marked = returned.pop("_total_marked", 0)       # metadado interno injetado pelo read_returned
    m = model_name or model.MODEL_TRANSLATE
    verbatim_n, ai_n, cost = 0, 0, 0.0
    touched, stopped = [], False
    for scene in sorted(returned):
        slot = returned[scene]
        touched.append(scene)
        if slot["verbatim"]:
            verbatim_n += _apply_verbatim(root, scene, slot["verbatim"])   # sempre $0
        if slot["nota"]:
            if max_usd is not None and cost >= max_usd:
                stopped = True
                continue                                  # teto: pula o caminho PAGO (verbatim ja entrou)
            note = "\n\n## REVISAO DO HUMANO (reescreva SO estes offsets seguindo a instrucao)\n" + \
                   "\n".join(f"- {off}: {ins}" for off, ins in slot["nota"])
            res = model.retranslate_offsets(root, scene, [o for o, _ in slot["nota"]],
                                            model=m, budget_tolerance=1.0, quality_note=note)
            if res.get("usage"):
                cost += model.cost_of(m, res["usage"])
            ai_n += len(slot["nota"])
    applied = verbatim_n + ai_n
    eff = round(applied / total_marked, 3) if total_marked else None
    rec = {"t": round(_time.time(), 3), "source": str(Path(csv_path).name),
           "total_marked": total_marked, "applied": applied,
           "verbatim": verbatim_n, "ai": ai_n,
           "effectiveness_rate": eff, "cost_usd": round(cost, 4)}
    try:
        with paths.qa_effectiveness(root).open("a", encoding="utf-8") as _fh:
            _fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except OSError:
        pass
    return {"verbatim": verbatim_n, "ai": ai_n, "scenes": len(touched),
            "cost_usd": round(cost, 4), "scenes_touched": touched, "stopped_budget": stopped,
            "total_marked": total_marked, "effectiveness_rate": eff}


def main():
    ap = argparse.ArgumentParser(description="Revisao humana por capitulo (export CSV marcado / apply do devolvido).")
    sub = ap.add_subparsers(dest="cmd", required=True)
    pe = sub.add_parser("export", help="gera o CSV marcado p/ revisao (capitulo, ou JOGO TODO se omitir)")
    pe.add_argument("project")
    pe.add_argument("chapter", nargs="?", default=None, help="capitulo (ex.: 11); OMITA p/ o jogo INTEIRO")
    pe.add_argument("--out", default=None)
    pe.add_argument("--csv", action="store_true", help="gera CSV cru (default: XLSX amigavel p/ o revisor)")
    pa = sub.add_parser("apply", help="aplica a revisao devolvida (verbatim + notas); le do INBOX por padrao")
    pa.add_argument("project")
    pa.add_argument("returned", nargs="?", default=None,
                    help="arquivo OU pasta devolvida; OMITA p/ ler do inbox (artifacts/qa_revisao/devolvido/)")
    pa.add_argument("--model", default=None)
    pa.add_argument("--max-usd", type=float, default=None, help="teto p/ o caminho de IA (notas); verbatim e $0")
    pw = sub.add_parser("width", help="GATE deterministico 'fora do balao': lista SO as linhas que estouram "
                                      "a largura visual (exit 1 se houver); loop ate zerar")
    pw.add_argument("project")
    pw.add_argument("chapter", nargs="?", default=None, help="capitulo; OMITA p/ o jogo INTEIRO")
    pw.add_argument("--out", default=None)
    pt = sub.add_parser("tester", help="relato in-game do TESTER (trecho do texto visto + print) -> localiza "
                                       "a linha e gera CORRIGIR no inbox (deterministico, $0, sem OCR/IA)")
    pt.add_argument("project")
    pt.add_argument("relato", nargs="?", default=None,
                    help="relato_tester.csv; OMITA p/ usar teste_ingame/relato_tester.csv")
    a = ap.parse_args()
    if a.cmd == "tester":
        relato = a.relato or str(paths.qa_tester(Path(a.project)) / "relato_tester.csv")
        if not Path(relato).is_file():
            sys.exit(f"[tester] relato nao encontrado: {relato} (preencha o template em teste_ingame/).")
        rows, ambiguous, missing = tester_to_review(a.project, relato)
        if rows:
            inbox = paths.qa_inbox(Path(a.project)); inbox.mkdir(parents=True, exist_ok=True)
            out = inbox / "relato_tester_resolvido.csv"
            write_csv(rows, out)
            print(f"[tester] {len(rows)} relato(s) localizado(s) -> {out} (rode: quality_review.py apply <projeto>)")
        else:
            print("[tester] nenhum relato localizado de forma unica.")
        for a_ in ambiguous:
            print(f"  AMBIGUO ({a_['n']}x) print={a_['print']} '{a_['texto_visto']}': {', '.join(a_['candidatos'])} "
                  f"-> desempate pelo print e edite o devolvido a mao")
        for m in missing:
            print(f"  NAO ACHADO print={m['print']} '{m['texto_visto']}' -> trecho maior/mais exato")
        sys.exit(0)
    if a.cmd == "width":
        rows = width_violations(a.project, a.chapter)
        scope = f"cap_{a.chapter}" if a.chapter else "all"
        label = f"cap.{a.chapter}" if a.chapter else "JOGO INTEIRO"
        if not rows:
            print(f"[width] {label}: OK — 0 linha(s) fora do balao (segmento <= {WIDTH_MAX} translit).")
            sys.exit(0)
        if a.out:
            out = a.out
        else:
            outbox = paths.qa_outbox(Path(a.project)); outbox.mkdir(parents=True, exist_ok=True)
            out = str(outbox / f"review_largura_{scope}.xlsx")
        write_xlsx(rows, out)                              # arquivo SO com as linhas problematicas
        print(f"[width] {label}: {len(rows)} linha(s) FORA DO BALAO (segmento > {WIDTH_MAX}) -> {out}")
        print("        Encurte (CORRIGIR + Correcao/Nota), rode 'apply', e re-rode 'width' ate zerar.")
        sys.exit(1)                                        # gate: exit !=0 alimenta o loop ate ficar limpo
    if a.cmd == "export":
        rows = export(a.project, a.chapter)
        scope = f"cap_{a.chapter}" if a.chapter else "all"
        ext = "csv" if a.csv else "xlsx"
        root_p = Path(a.project)
        outbox = paths.qa_outbox(root_p)
        inbox  = paths.qa_inbox(root_p)
        tester = paths.qa_tester(root_p)
        for d in (outbox, inbox, tester):          # garante estrutura completa antes de qualquer arquivo
            d.mkdir(parents=True, exist_ok=True)
        if a.out:
            out = a.out
        else:
            out = str(outbox / f"review_{scope}.{ext}")
        (write_csv if a.csv else write_xlsx)(rows, out)
        marked = sum(1 for r in rows if r["revisar"])
        label = f"cap.{a.chapter}" if a.chapter else "JOGO INTEIRO"
        print(f"[export] {label}: {len(rows)} linha(s) -> {out}")
        print(f"         {marked} marcada(s) p/ avaliar; abra no Excel/LibreOffice, filtre a coluna "
              f"'Revisar', preencha 'Correcao' (texto certo) ou 'Nota' (instrucao) e devolva em "
              f"{inbox} (ou em {tester}/relato_tester.csv para relatos in-game).")
        sys.exit(0)
    files = returned_files(Path(a.project), a.returned)
    if not files:
        print(f"[apply] nada a aplicar — nenhum .xlsx/.csv devolvido em "
              f"{a.returned or paths.qa_inbox(Path(a.project))}.")
        sys.exit(0)
    tot = {"verbatim": 0, "ai": 0, "cost": 0.0, "scenes": set(), "stopped": False}
    for f in files:
        print(f"[apply] processando revisao devolvida: {f}")
        r = apply(a.project, f, model_name=a.model, max_usd=a.max_usd)
        tot["verbatim"] += r["verbatim"]; tot["ai"] += r["ai"]; tot["cost"] += r["cost_usd"]
        tot["scenes"].update(r["scenes_touched"]); tot["stopped"] = tot["stopped"] or r.get("stopped_budget")
    print(f"[apply] TOTAL: verbatim={tot['verbatim']} (0 IA) | nota+IA={tot['ai']} (~${tot['cost']:.4f}) "
          f"| cenas tocadas={len(tot['scenes'])}")
    if tot["stopped"]:
        print(f"[apply] teto de ${a.max_usd:.2f} atingido — algumas notas (IA) nao foram processadas; "
              "verbatim entrou tudo. Re-rode com mais orcamento p/ as notas restantes.")
    print("Proximos passos: verify_chapter de cada cap. tocado (round-trip/charset) + state_index --rebuild.")
    print(f"  cenas: {', '.join(sorted(tot['scenes']))}")
    sys.exit(0)


if __name__ == "__main__":
    main()
