#!/usr/bin/env python3
"""
kb_fetch.py — busca e normaliza fontes de KB research p/ texto plano, SEM LLM e sem chave de API.

Parte 1 do pipeline hibrido de KB (onboarding de baixo custo): o humano passa qualquer fonte (URL de wiki/site, PDF,
.docx, .xlsx, arquivo local .txt/.md) e este script normaliza tudo pra texto simples, cacheado em
artifacts/research_cache/<hash>.md. O kb_build_ollama.py le esse cache sem saber se a origem era
web ou arquivo local — a normalizacao acontece so aqui, uma vez.

Zero raciocinio, zero chamada de modelo: so download/leitura + extracao deterministica. O ganho de
tokens do onboarding de baixo custo vem de tirar essa leitura/extracao bruta da sessao Claude (ver skill 03, Fase 1A).

Deps: URL usa so stdlib (urllib + html.parser — mesmo padrao do ollama_client.py, sem lib nova).
PDF/.docx exigem pdfplumber/python-docx (requirements-kb.txt, OPCIONAL); import preguicoso com
RuntimeError claro se faltar, mesmo padrao do backend 'api' faltando o SDK anthropic (llm_client.py).
.xlsx usa openpyxl (ja em requirements-dev.txt, sem lazy-import).

Uso:  python kb_fetch.py <projeto> <fonte_1> [fonte_2] ...
      <fonte> = URL http(s):// ou caminho local (.pdf/.docx/.xlsx/.txt/.md/outro).
"""
from __future__ import annotations

import hashlib
import re
import sys
import urllib.error
import urllib.request
from datetime import UTC, datetime
from html.parser import HTMLParser
from pathlib import Path

_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))
import paths  # noqa: E402  (paths.py: fonte unica do contrato de caminhos de artefato)

_MAX_BYTES = 10_000_000   # teto por fonte (10 MB) — nao deixa 1 fonte gigante travar o fetch
_TIMEOUT = 20             # segundos por requisicao HTTP
_USER_AGENT = "translation-cognition-framework/kb_fetch (+local research tool)"

_URL_RE = re.compile(r"^https?://", re.I)
_BLOCK_TAGS = {"p", "li", "h1", "h2", "h3", "h4", "h5", "h6", "br", "div", "tr", "title"}
# Remove <script>/<style> ANTES do HTMLParser ver o HTML. Nao da p/ confiar em start/end tag do
# proprio parser p/ isso: o stdlib html.parser trata <script>/<style> como CDATA internamente (modo
# especial, independente de handle_starttag/endtag) -- se o fechamento nunca chega (HTML malformado
# real, ou cortado pelo teto _MAX_BYTES bem no meio do bloco), o parser fica esperando o terminador
# p/ SEMPRE e todo o texto depois na pagina e descartado em silencio (bug real, pego por teste).
# Fix em 2 passos: 1) remove pares FECHADOS; 2) qualquer <script>/<style> que sobrar depois disso e,
# por definicao, orfao (sem fechamento) -- remove so a tag de abertura, pra nunca deixar o parser
# entrar no modo CDATA sem saida.
_SCRIPT_STYLE_RE = re.compile(r"<(script|style)\b[^>]*>.*?</\1\s*>", re.I | re.S)
_ORPHAN_SCRIPT_STYLE_OPEN_RE = re.compile(r"<(?:script|style)\b[^>]*>", re.I)


class _TextExtractor(HTMLParser):
    """Extrai texto de HTML cru sem lib nova (stdlib). Extracao crua, nao visual — insere quebra
    de linha nos blocos comuns. P/ fontes com muito boilerplate de layout, preferir salvar a pagina
    como .txt manualmente em vez de passar a URL."""

    def __init__(self):
        super().__init__()
        self._chunks: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag in _BLOCK_TAGS:
            self._chunks.append("\n")

    def handle_data(self, data):
        t = data.strip()
        if t:
            self._chunks.append(t)

    def text(self) -> str:
        joined = " ".join(self._chunks)
        joined = re.sub(r"[ \t]*\n[ \t]*", "\n", joined)
        joined = re.sub(r"\n{3,}", "\n\n", joined)
        return joined.strip()


def _extract_url(url: str) -> tuple[str, bool]:
    req = urllib.request.Request(url, headers={"User-Agent": _USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=_TIMEOUT) as resp:  # nosec B310 - URL fornecida explicitamente pelo humano (skill 03)
            raw = resp.read(_MAX_BYTES + 1)
    except urllib.error.URLError as e:
        raise RuntimeError(f"falha ao buscar {url}: {e}") from e
    truncado = len(raw) > _MAX_BYTES
    if truncado:
        raw = raw[:_MAX_BYTES]
    html = raw.decode("utf-8", errors="replace")
    html = _SCRIPT_STYLE_RE.sub(" ", html)
    html = _ORPHAN_SCRIPT_STYLE_OPEN_RE.sub(" ", html)
    parser = _TextExtractor()
    parser.feed(html)
    return parser.text(), truncado


def _extract_pdf(path: Path) -> str:
    try:
        import pdfplumber
    except ImportError as e:
        raise RuntimeError(
            "fonte PDF requer o pacote 'pdfplumber' (pip install -r requirements-kb.txt)."
        ) from e
    texts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    return "\n\n".join(texts)


def _extract_docx(path: Path) -> str:
    try:
        import docx
    except ImportError as e:
        raise RuntimeError(
            "fonte .docx requer o pacote 'python-docx' (pip install -r requirements-kb.txt)."
        ) from e
    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"# {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(",".join(cells))
    return "\n".join(lines)


def _extract_local_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def fetch_one(source: str) -> dict:
    """Busca UMA fonte (URL http(s) ou caminho local) e retorna texto normalizado, SEM tocar
    cache em disco (pura o bastante p/ ser testavel isolada). {"fonte", "tipo", "texto", "truncado"}.
    Levanta RuntimeError com mensagem clara p/ dependencia ausente ou falha de rede/arquivo."""
    if _URL_RE.match(source):
        texto, truncado = _extract_url(source)
        return {"fonte": source, "tipo": "url", "texto": texto, "truncado": truncado}
    path = Path(source)
    if not path.is_file():
        raise RuntimeError(f"fonte local nao encontrada: {source}")
    ext = path.suffix.lower()
    if ext == ".pdf":
        texto = _extract_pdf(path)
    elif ext == ".docx":
        texto = _extract_docx(path)
    elif ext == ".xlsx":
        texto = _extract_xlsx(path)
    else:
        texto = _extract_local_text(path)
    return {"fonte": source, "tipo": ext.lstrip(".") or "texto", "texto": texto, "truncado": False}


def _hash_of(source: str) -> str:
    return hashlib.sha1(source.encode("utf-8"), usedforsecurity=False).hexdigest()[:16]


def fetch_and_cache(root, source: str, *, fetched_em: str, found_by: str = "ia") -> Path:
    """fetch_one + grava o cache em artifacts/research_cache/<hash>.md com front-matter (fonte,
    tipo, timestamp, truncado, encontrada_por) — e de onde kb_build_ollama.py tira a citacao de
    fonte (e QUEM achou) por entrada do research_log.md, mesmo sem saber se a origem era web ou
    local. `found_by` restaura a distincao IA/usuario que a skill 03 (Fase 1A/1B) precisa p/ tier —
    sem isso toda fonte vira so "achada pelo Ollama", perdendo a proveniencia da pesquisa humana.
    hash = sha1(fonte), nao do conteudo -> mesmo nome sempre p/ a mesma fonte (re-fetch idempotente)."""
    root = Path(root)
    r = fetch_one(source)
    out = paths.research_cache(root, _hash_of(source))
    out.parent.mkdir(parents=True, exist_ok=True)
    front = (
        "---\n"
        f"fonte: {r['fonte']}\n"
        f"tipo: {r['tipo']}\n"
        f"fetched_em: {fetched_em}\n"
        f"truncado: {'true' if r['truncado'] else 'false'}\n"
        f"encontrada_por: {found_by}\n"
        "---\n"
    )
    out.write_text(front + r["texto"], encoding="utf-8")
    return out


def main():
    import argparse
    ap = argparse.ArgumentParser(description="Busca e normaliza fontes de KB research (sem LLM).")
    ap.add_argument("project")
    ap.add_argument("sources", nargs="+", help="URL(s) http(s):// ou caminho(s) local(is)")
    ap.add_argument("--found-por", choices=["ia", "usuario"], default="ia",
                    help="quem achou estas fontes (proveniencia p/ tiering da skill 03, Fase 1A/1B)")
    a = ap.parse_args()
    fetched_em = datetime.now(UTC).isoformat()
    for src in a.sources:
        try:
            out = fetch_and_cache(a.project, src, fetched_em=fetched_em, found_by=a.found_por)
            print(f"[kb_fetch] {src} -> {out}")
        except RuntimeError as e:
            print(f"[kb_fetch] FALHOU: {src} -> {e}")
            sys.exit(1)


if __name__ == "__main__":
    main()
